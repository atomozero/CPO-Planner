# documents/views.py
import io
import openpyxl
import os
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from decimal import Decimal
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse, Http404, FileResponse
from django.urls import reverse, reverse_lazy
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, FormView, View
)
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.contenttypes.models import ContentType
from django.contrib import messages
from django.db.models import Q, Sum
from django.utils import timezone
from django.utils.translation import gettext_lazy as _
from django.core.exceptions import PermissionDenied
from django.conf import settings

# Importa dai modelli corretti
from cpo_core.models.project import Project
from cpo_core.models.subproject import SubProject
from infrastructure.models import Municipality
from cpo_planner.projects.models import ChargingStation
from .models import (
    ProjectDocument, 
    DocumentTemplate,
    DocumentNote,
    DocumentTask,
    DocumentCategory,
    Document
)
from .forms import (
    ProjectDocumentForm, 
    DocumentNoteForm, 
    DocumentTaskForm,
    DocumentFilterForm,
    DocumentCategoryForm
)

class ProjectDocumentListView(LoginRequiredMixin, ListView):
    """Vista per elencare i documenti dei progetti"""
    model = ProjectDocument
    template_name = 'documents/project_document_list.html'
    context_object_name = 'documents'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Logica di filtraggio basata sull'entità
        project_id = self.kwargs.get('project_id')
        if project_id:
            project = get_object_or_404(Project, id=project_id)
            queryset = queryset.filter(project=project)
        
        # Applica filtri dal form
        form = DocumentFilterForm(self.request.GET)
        if form.is_valid():
            # Filtri specifici
            if form.cleaned_data.get('search'):
                queryset = queryset.filter(
                    Q(name__icontains=form.cleaned_data['search']) | 
                    Q(description__icontains=form.cleaned_data['search'])
                )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi contesto per il progetto correlato
        project_id = self.kwargs.get('project_id')
        if project_id:
            context['project'] = get_object_or_404(Project, id=project_id)
        
        return context

class ProjectDocumentCreateView(LoginRequiredMixin, CreateView):
    """Vista per creare un nuovo documento di progetto"""
    model = ProjectDocument
    form_class = ProjectDocumentForm
    template_name = 'documents/project_document_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        
        # Passa il progetto se specificato
        project_id = self.kwargs.get('project_id')
        if project_id:
            kwargs['project'] = get_object_or_404(Project, id=project_id)
        
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi il progetto al contesto se specificato
        project_id = self.kwargs.get('project_id')
        if project_id:
            context['project'] = get_object_or_404(Project, id=project_id)
        
        return context
    
    def form_valid(self, form):
        """
        Imposta automaticamente il progetto se non è già stato fatto
        """
        project_id = self.kwargs.get('project_id')
        if project_id and not form.instance.project:
            form.instance.project = get_object_or_404(Project, id=project_id)
        
        # Imposta l'utente corrente come creatore
        form.instance.created_by = self.request.user
        
        return super().form_valid(form)
    
    def get_success_url(self):
        messages.success(self.request, _('Documento creato con successo.'))
        
        # Reindirizza al dettaglio del documento o alla lista documenti del progetto
        if self.object.project:
            return reverse('documents:project_document_list', 
                           kwargs={'project_id': self.object.project.id})
        return reverse('documents:project_document_list')

class ProjectDocumentDetailView(LoginRequiredMixin, DetailView):
    """Vista dettaglio di un documento di progetto"""
    model = ProjectDocument
    template_name = 'documents/project_document_detail.html'
    context_object_name = 'document'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi note e task al contesto
        context['notes'] = self.object.notes.order_by('-created_at')
        context['tasks'] = self.object.tasks.order_by('-due_date')
        
        # Form per nuove note e task
        context['note_form'] = DocumentNoteForm()
        context['task_form'] = DocumentTaskForm()
        
        return context

class ProjectDocumentDownloadView(LoginRequiredMixin, DetailView):
    """Vista per scaricare un documento di progetto"""
    model = ProjectDocument
    
    def get(self, request, *args, **kwargs):
        document = self.get_object()
        
        # Controlla l'esistenza del file
        if not document.file:
            messages.error(request, _('Il file non esiste.'))
            return redirect('documents:project_document_detail', pk=document.pk)
        
        # Scarica il file
        try:
            response = HttpResponse(document.file.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename={document.file.name}'
            return response
        except Exception as e:
            messages.error(request, _('Errore durante il download: {}').format(str(e)))
            return redirect('documents:project_document_detail', pk=document.pk)

class DocumentTemplateListView(LoginRequiredMixin, ListView):
    """Vista per elencare i template di documenti"""
    model = DocumentTemplate
    template_name = 'documents/document_template_list.html'
    context_object_name = 'templates'

class DocumentTemplateCreateView(LoginRequiredMixin, CreateView):
    """Vista per creare un nuovo template di documento"""
    model = DocumentTemplate
    fields = ['name', 'description', 'file_type', 'template_file']
    template_name = 'documents/document_template_form.html'
    success_url = reverse_lazy('documents:template_list')

def generate_project_business_plan(request, project_id):
    """
    Genera un business plan PDF per un progetto specifico
    """
    project = get_object_or_404(Project, id=project_id)
    
    try:
        # Preparazione del buffer per il PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Titolo del documento
        story.append(Paragraph(f"Business Plan - {project.name}", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Informazioni di base del progetto
        story.append(Paragraph("Informazioni Progetto", styles['Heading2']))
        project_info = [
            ['Nome Progetto', project.name],
            ['Regione', project.region],
            ['Data Inizio', project.start_date.strftime('%d/%m/%Y')],
            ['Data Completamento Prevista', project.expected_completion_date.strftime('%d/%m/%Y')],
        ]
        project_info_table = Table(project_info, colWidths=[150, 300])
        project_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.grey),
            ('TEXTCOLOR', (0,0), (0,-1), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('BACKGROUND', (1,0), (1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(project_info_table)
        story.append(Spacer(1, 12))
        
        # Analisi Finanziaria
        story.append(Paragraph("Analisi Finanziaria", styles['Heading2']))
        financial_data = [
            ['Metrica', 'Valore'],
            ['Budget Totale', f'€ {project.total_budget:,.2f}'],
            ['Ricavi Attesi', f'€ {project.total_expected_revenue:,.2f}'],
            ['ROI Totale', f'{project.total_roi:.2f}%'],
        ]
        financial_table = Table(financial_data, colWidths=[200, 250])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('BACKGROUND', (1,1), (1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(financial_table)
        story.append(Spacer(1, 12))
        
        # Sotto-Progetti
        story.append(Paragraph("Sotto-Progetti", styles['Heading2']))
        sub_projects = project.subproject_set.all()
        sub_project_data = [['Nome', 'Comune', 'Budget', 'Ricavi Attesi']]
        for sp in sub_projects:
            sub_project_data.append([
                sp.name, 
                str(sp.municipality), 
                f'€ {sp.budget:,.2f}', 
                f'€ {sp.expected_revenue:,.2f}'
            ])
        
        sub_project_table = Table(sub_project_data, colWidths=[150, 150, 100, 100])
        sub_project_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(sub_project_table)
        story.append(Spacer(1, 12))
        
        # Stazioni di Ricarica
        story.append(Paragraph("Stazioni di Ricarica", styles['Heading2']))
        stations = ChargingStation.objects.filter(sub_project__project=project)
        stations_data = [['Identificatore', 'Indirizzo', 'Tipo Ricarica', 'Potenza Max']]
        for station in stations:
            stations_data.append([
                station.identifier, 
                station.address, 
                station.charging_type, 
                f'{station.max_power} kW'
            ])
        
        stations_table = Table(stations_data, colWidths=[100, 200, 100, 100])
        stations_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(stations_table)
        
        # Genera il PDF
        doc.build(story)
        buffer.seek(0)
        
        # Salva il documento
        project_document = ProjectDocument.objects.create(
            project=project,
            name=f"Business Plan {project.name}",
            description="Business plan generato automaticamente",
            file=buffer
        )
        
        messages.success(request, _('Business Plan generato con successo.'))
        return FileResponse(buffer, as_attachment=True, 
                            filename=f"business_plan_{project.name}.pdf")
    
    except Exception as e:
        messages.error(request, _('Errore durante la generazione del Business Plan: {}').format(str(e)))
        return redirect('projects:project_detail', pk=project_id)

def add_document_note(request, document_id):
    """
    Aggiunge una nota a un documento
    """
    document = get_object_or_404(ProjectDocument, id=document_id)
    
    if request.method == 'POST':
        form = DocumentNoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.document = document
            note.created_by = request.user
            note.save()
            messages.success(request, _('Nota aggiunta con successo.'))
        else:
            messages.error(request, _('Errore nell\'aggiunta della nota.'))
    
    return redirect('documents:project_document_detail', pk=document_id)

def add_document_task(request, document_id):
    """
    Aggiunge un'attività a un documento
    """
    document = get_object_or_404(ProjectDocument, id=document_id)
    
    if request.method == 'POST':
        form = DocumentTaskForm(request.POST)
        if form.is_valid():
            task = form.save(commit=False)
            task.document = document
            task.created_by = request.user
            task.save()
            messages.success(request, _('Attività aggiunta con successo.'))
        else:
            messages.error(request, _('Errore nell\'aggiunta dell\'attività.'))
    
    return redirect('documents:project_document_detail', pk=document_id)
    
def update_document_task_status(request, task_id):
    """
    Aggiorna lo stato di un'attività di un documento
    """
    task = get_object_or_404(DocumentTask, id=task_id)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        
        if new_status:
            task.status = new_status
            if new_status == 'completed':
                task.completed_at = timezone.now()
            task.save()
            
            messages.success(request, _('Stato attività aggiornato con successo.'))
            return redirect('documents:project_document_detail', pk=task.document.id)
        
        messages.error(request, _('Stato non valido.'))
    
    return redirect('documents:project_document_detail', pk=task.document.id)

def generate_charging_station_report(request, station_id):
    """
    Genera un rapporto dettagliato per una singola stazione di ricarica
    """
    station = get_object_or_404(ChargingStation, id=station_id)
    
    try:
        # Preparazione del buffer per il PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Titolo del documento
        story.append(Paragraph(f"Rapporto Stazione di Ricarica: {station.identifier}", styles['Title']))
        story.append(Spacer(1, 12))
        
        # Informazioni Base della Stazione
        station_info = [
            ['Dettaglio', 'Valore'],
            ['Identificatore', station.identifier],
            ['Indirizzo', station.address],
            ['Coordinate', f'{station.latitude}, {station.longitude}'],
            ['Tipo di Ricarica', station.charging_type],
            ['Potenza Massima', f'{station.max_power} kW'],
        ]
        station_info_table = Table(station_info, colWidths=[200, 250])
        station_info_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.grey),
            ('TEXTCOLOR', (0,0), (0,-1), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('BACKGROUND', (1,0), (1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(station_info_table)
        story.append(Spacer(1, 12))
        
        # Analisi Finanziaria
        financial_data = [
            ['Metrica Finanziaria', 'Valore'],
            ['Costo Installazione', f'€ {station.installation_cost:,.2f}'],
            ['Costo Allaccio', f'€ {station.connection_cost:,.2f}'],
            ['Ricavi Mensili Stimati', f'€ {station.estimated_monthly_revenue:,.2f}'],
        ]
        
        # Calcolo metriche annuali
        annual_metrics = station.calculate_annual_metrics()
        financial_data.extend([
            ['Ricavi Annuali', f'€ {annual_metrics["annual_revenue"]:,.2f}'],
            ['Costi Annuali', f'€ {annual_metrics["annual_costs"]:,.2f}'],
            ['Profitto Annuale', f'€ {annual_metrics["annual_profit"]:,.2f}'],
        ])
        
        financial_table = Table(financial_data, colWidths=[200, 250])
        financial_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.grey),
            ('TEXTCOLOR', (0,0), (0,-1), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('BACKGROUND', (1,0), (1,-1), colors.beige),
            ('GRID', (0,0), (-1,-1), 1, colors.black)
        ]))
        story.append(financial_table)
        story.append(Spacer(1, 12))
        
        # Impianto Fotovoltaico
        if station.has_photovoltaic_system:
            pv_data = [
                ['Dettaglio Impianto Fotovoltaico', 'Valore'],
                ['Presente', 'Sì'],
                ['Capacità', f'{station.photovoltaic_capacity} kWp'],
            ]
            pv_table = Table(pv_data, colWidths=[200, 250])
            pv_table.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,-1), colors.grey),
                ('TEXTCOLOR', (0,0), (0,-1), colors.whitesmoke),
                ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE', (0,0), (-1,0), 12),
                ('BOTTOMPADDING', (0,0), (-1,-1), 12),
                ('BACKGROUND', (1,0), (1,-1), colors.beige),
                ('GRID', (0,0), (-1,-1), 1, colors.black)
            ]))
            story.append(pv_table)
        else:
            story.append(Paragraph("Nessun Impianto Fotovoltaico Presente", styles['Normal']))
        
        # Genera il PDF
        doc.build(story)
        buffer.seek(0)
        
        # Salva il documento
        station_document = ProjectDocument.objects.create(
            project=station.sub_project.project,
            name=f"Rapporto Stazione {station.identifier}",
            description="Rapporto dettagliato stazione di ricarica",
            file=buffer
        )
        
        messages.success(request, _('Rapporto Stazione generato con successo.'))
        return FileResponse(buffer, as_attachment=True, 
                            filename=f"station_report_{station.identifier}.pdf")
    
    except Exception as e:
        messages.error(request, _('Errore durante la generazione del Rapporto Stazione: {}').format(str(e)))
        return redirect('projects:charging_station_detail', pk=station_id)

def export_project_financial_summary(request, project_id):
    """
    Esporta un riepilogo finanziario dettagliato del progetto in formato Excel
    """
    project = get_object_or_404(Project, id=project_id)
    
    try:
        # Preparazione del workbook Excel
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # Foglio Informazioni Progetto
        sheet_info = workbook.active
        sheet_info.title = "Informazioni Progetto"
        sheet_info.append(["Nome Progetto", project.name])
        sheet_info.append(["Regione", project.region])
        sheet_info.append(["Data Inizio", project.start_date])
        sheet_info.append(["Data Completamento Prevista", project.expected_completion_date])
        
        # Foglio Analisi Finanziaria
        sheet_financial = workbook.create_sheet(title="Analisi Finanziaria")
        sheet_financial.append([
            "Metrica", "Valore"
        ])
        sheet_financial.append(["Budget Totale", project.total_budget])
        sheet_financial.append(["Ricavi Attesi", project.total_expected_revenue])
        sheet_financial.append(["ROI Totale", project.total_roi])
        
        # Foglio Sotto-Progetti
        sheet_subprojects = workbook.create_sheet(title="Sotto-Progetti")
        sheet_subprojects.append([
            "Nome", "Comune", "Budget", "Ricavi Attesi", "ROI"
        ])
        for sp in project.subproject_set.all():
            sheet_subprojects.append([
                sp.name, 
                str(sp.municipality), 
                sp.budget, 
                sp.expected_revenue,
                sp.roi
            ])
        
        # Foglio Stazioni di Ricarica
        sheet_stations = workbook.create_sheet(title="Stazioni di Ricarica")
        sheet_stations.append([
            "Identificatore", "Indirizzo", "Tipo Ricarica", 
            "Potenza Max", "Costo Installazione", "Ricavi Mensili"
        ])
        stations = ChargingStation.objects.filter(sub_project__project=project)
        for station in stations:
            sheet_stations.append([
                station.identifier,
                station.address,
                station.charging_type,
                station.max_power,
                station.installation_cost,
                station.estimated_monthly_revenue
            ])
        
        # Salva il workbook
        workbook.save(buffer)
        buffer.seek(0)
        
        # Salva il documento
        project_document = ProjectDocument.objects.create(
            project=project,
            name=f"Riepilogo Finanziario {project.name}",
            description="Riepilogo finanziario dettagliato del progetto",
            file=buffer
        )
        
        messages.success(request, _('Riepilogo Finanziario generato con successo.'))
        return FileResponse(buffer, as_attachment=True, 
                            filename=f"financial_summary_{project.name}.xlsx")
    
    except Exception as e:
        messages.error(request, _('Errore durante la generazione del Riepilogo Finanziario: {}').format(str(e)))
        return redirect('projects:project_detail', pk=project_id)


# Eventuali configurazioni aggiuntive
def get_municipality_projects_summary(request, municipality_id):
    """
    Genera un rapporto riassuntivo dei progetti per un comune specifico
    """
    municipality = get_object_or_404(Municipality, id=municipality_id)
    
    # Recupera i sotto-progetti per questo comune
    sub_projects = SubProject.objects.filter(municipality=municipality)
    
    # Calcola metriche aggregate
    total_budget = sub_projects.aggregate(Sum('budget'))['budget__sum'] or Decimal('0.00')
    total_expected_revenue = sub_projects.aggregate(Sum('expected_revenue'))['expected_revenue__sum'] or Decimal('0.00')
    
    # Recupera le stazioni di ricarica
    charging_stations = ChargingStation.objects.filter(sub_project__municipality=municipality)
    
    context = {
        'municipality': municipality,
        'sub_projects': sub_projects,
        'total_budget': total_budget,
        'total_expected_revenue': total_expected_revenue,
        'charging_stations': charging_stations,
    }
    
    return render(request, 'documents/municipality_projects_summary.html', context)

class DocumentListView(LoginRequiredMixin, ListView):
    """Vista per elencare i documenti per una specifica entità"""
    model = ProjectDocument
    template_name = 'documents/document_list.html'
    context_object_name = 'documents'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Filtra per tipo di entità e ID
        entity_type = self.kwargs.get('entity_type')
        entity_id = self.kwargs.get('entity_id')
        
        if entity_type == 'project':
            queryset = queryset.filter(project_id=entity_id)
        elif entity_type == 'subproject':
            queryset = queryset.filter(project__subproject__id=entity_id)
        elif entity_type == 'chargingstation':
            queryset = queryset.filter(project__subproject__chargingstation__id=entity_id)
            
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi informazioni sull'entità
        entity_type = self.kwargs.get('entity_type')
        entity_id = self.kwargs.get('entity_id')
        
        if entity_type == 'project':
            context['entity'] = get_object_or_404(Project, id=entity_id)
        elif entity_type == 'subproject':
            context['entity'] = get_object_or_404(SubProject, id=entity_id)
        elif entity_type == 'chargingstation':
            context['entity'] = get_object_or_404(ChargingStation, id=entity_id)
            
        context['entity_type'] = entity_type
        return context
    
class DocumentDetailView(LoginRequiredMixin, DetailView):
    """Vista dettaglio di un documento"""
    model = ProjectDocument
    template_name = 'documents/document_detail.html'
    context_object_name = 'document'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi note e task al contesto
        context['notes'] = self.object.notes.order_by('-created_at')
        context['tasks'] = self.object.tasks.order_by('-due_date')
        
        # Form per nuove note e task
        context['note_form'] = DocumentNoteForm()
        context['task_form'] = DocumentTaskForm()
        
        return context

class DocumentCreateView(LoginRequiredMixin, CreateView):
    """Vista per creare un nuovo documento"""
    model = ProjectDocument
    form_class = ProjectDocumentForm
    template_name = 'documents/document_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        
        # Passa l'entità se specificata
        entity_type = self.kwargs.get('entity_type')
        entity_id = self.kwargs.get('entity_id')
        
        if entity_type and entity_id:
            if entity_type == 'project':
                kwargs['entity_obj'] = get_object_or_404(Project, id=entity_id)
                kwargs['entity_type'] = 'project'
            elif entity_type == 'subproject':
                kwargs['entity_obj'] = get_object_or_404(SubProject, id=entity_id)
                kwargs['entity_type'] = 'subproject'
            elif entity_type == 'chargingstation':
                kwargs['entity_obj'] = get_object_or_404(ChargingStation, id=entity_id)
                kwargs['entity_type'] = 'chargingstation'
        
        # Passa l'utente attuale
        kwargs['user'] = self.request.user
        
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Aggiungi l'entità al contesto se specificata
        entity_type = self.kwargs.get('entity_type')
        entity_id = self.kwargs.get('entity_id')
        
        if entity_type and entity_id:
            if entity_type == 'project':
                context['entity'] = get_object_or_404(Project, id=entity_id)
            elif entity_type == 'subproject':
                context['entity'] = get_object_or_404(SubProject, id=entity_id)
            elif entity_type == 'chargingstation':
                context['entity'] = get_object_or_404(ChargingStation, id=entity_id)
                
            context['entity_type'] = entity_type
        
        return context
    
    def get_success_url(self):
        messages.success(self.request, _('Documento creato con successo.'))
        
        # Reindirizza alla lista documenti dell'entità o alla lista generale
        entity_type = self.kwargs.get('entity_type')
        entity_id = self.kwargs.get('entity_id')
        
        if entity_type and entity_id:
            return reverse('documents:document_list_entity', 
                          kwargs={'entity_type': entity_type, 'entity_id': entity_id})
        
        return reverse('documents:document_list')

class DocumentUpdateView(LoginRequiredMixin, UpdateView):
    """Vista per aggiornare un documento esistente"""
    model = ProjectDocument
    form_class = ProjectDocumentForm
    template_name = 'documents/document_form.html'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        # Passa l'utente corrente al form
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        return context
    
    def get_success_url(self):
        messages.success(self.request, _('Documento aggiornato con successo.'))
        return reverse('documents:document_detail', kwargs={'pk': self.object.pk})
    
class DocumentDeleteView(LoginRequiredMixin, View):
    """Vista per eliminare un documento, supporta sia Document che ProjectDocument"""
    template_name = 'documents/document_confirm_delete.html'
    
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        document = self.get_document(pk)
        
        if not document:
            messages.error(request, _('Documento non trovato.'))
            return redirect('documents:document_list')
        
        return render(request, self.template_name, {'document': document})
    
    def post(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        document = self.get_document(pk)
        
        if not document:
            messages.error(request, _('Documento non trovato.'))
            return redirect('documents:document_list')
        
        project = None
        if hasattr(document, 'project') and document.project:
            project = document.project
        
        # Elimina il documento
        try:
            document.delete()
            messages.success(request, _('Documento eliminato con successo.'))
            
            # Reindirizza in base all'esistenza di un progetto associato
            if project:
                return redirect('documents:document_list_entity', 
                                entity_type='project', 
                                entity_id=project.id)
            else:
                return redirect('documents:document_list')
                
        except Exception as e:
            messages.error(request, _('Errore durante l\'eliminazione: {}').format(str(e)))
            return redirect('documents:document_list')
    
    def get_document(self, pk):
        """Trova il documento usando diversi modelli"""
        # Prima prova con il modello ProjectDocument da projects
        try:
            from cpo_planner.projects.models.document import ProjectDocument as ProjectsProjectDocument
            return ProjectsProjectDocument.objects.get(pk=pk)
        except:
            try:
                # Prova con il modello ProjectDocument da documents
                return ProjectDocument.objects.get(pk=pk)
            except:
                try:
                    # Prova con Document generico
                    return Document.objects.get(pk=pk)
                except:
                    return None

class DocumentDownloadView(LoginRequiredMixin, View):
    """Vista per scaricare un documento, supporta sia Document che ProjectDocument"""
    
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        
        # Prima prova con il modello ProjectDocument
        try:
            from cpo_planner.projects.models.document import ProjectDocument as ProjectsProjectDocument
            document = ProjectsProjectDocument.objects.get(pk=pk)
        except:
            try:
                # Prova con il modello ProjectDocument da documents
                document = ProjectDocument.objects.get(pk=pk)
            except:
                try:
                    # Prova con Document generico
                    document = Document.objects.get(pk=pk)
                except:
                    messages.error(request, _('Documento non trovato.'))
                    return redirect('documents:document_list')
        
        # Controlla l'esistenza del file
        if not document.file:
            messages.error(request, _('Il file non esiste.'))
            if hasattr(document, 'project'):
                return redirect('projects:project_detail', pk=document.project.pk)
            else:
                return redirect('documents:document_list')
        
        # Scarica il file
        try:
            file_path = document.file.path
            file_name = os.path.basename(document.file.name)
            
            # Usa FileResponse per gestire correttamente il file
            response = FileResponse(
                open(file_path, 'rb'),
                content_type='application/octet-stream',
                as_attachment=True,
                filename=file_name
            )
            return response
        except Exception as e:
            messages.error(request, _('Errore durante il download: {}').format(str(e)))
            if hasattr(document, 'project'):
                return redirect('projects:project_detail', pk=document.project.pk)
            else:
                return redirect('documents:document_list')
        
class DocumentPreviewView(LoginRequiredMixin, View):
    """Vista per visualizzare l'anteprima di un documento, supporta sia Document che ProjectDocument"""
    template_name = 'documents/document_preview.html'
    
    def get(self, request, *args, **kwargs):
        pk = kwargs.get('pk')
        
        # Prima prova con il modello ProjectDocument da projects
        try:
            from cpo_planner.projects.models.document import ProjectDocument as ProjectsProjectDocument
            document = ProjectsProjectDocument.objects.get(pk=pk)
        except:
            try:
                # Prova con il modello ProjectDocument da documents
                document = ProjectDocument.objects.get(pk=pk)
            except:
                try:
                    # Prova con Document generico
                    document = Document.objects.get(pk=pk)
                except:
                    messages.error(request, _('Documento non trovato.'))
                    return redirect('documents:document_list')
        
        # Controlla l'esistenza del file
        if not document.file:
            messages.error(request, _('Il file non esiste.'))
            if hasattr(document, 'project'):
                return redirect('projects:project_detail', pk=document.project.pk)
            else:
                return redirect('documents:document_list')
        
        # Per i documenti PDF e immagini, restituisci il file per la visualizzazione in-browser
        file_ext = os.path.splitext(document.file.name)[1].lower()
        if file_ext in ['.pdf', '.jpg', '.jpeg', '.png', '.gif']:
            try:
                response = FileResponse(
                    open(document.file.path, 'rb'),
                    content_type=self.get_content_type(file_ext)
                )
                return response
            except Exception as e:
                messages.error(request, _('Errore durante l\'anteprima: {}').format(str(e)))
                if hasattr(document, 'project'):
                    return redirect('projects:project_detail', pk=document.project.pk)
                else:
                    return redirect('documents:document_list')
        
        # Per altri tipi di file, reindirizza al download
        return redirect('documents:document_download', pk=document.pk)
    
    def get_content_type(self, ext):
        """Determina il content type in base all'estensione del file"""
        if ext == '.pdf':
            return 'application/pdf'
        elif ext in ['.jpg', '.jpeg']:
            return 'image/jpeg'
        elif ext == '.png':
            return 'image/png'
        elif ext == '.gif':
            return 'image/gif'
        else:
            return 'application/octet-stream'

def update_document_status(request, pk):
    """
    Aggiorna lo stato di un documento
    """
    document = get_object_or_404(ProjectDocument, id=pk)
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        
        if new_status:
            document.status = new_status
            document.save()
            
            messages.success(request, _('Stato documento aggiornato con successo.'))
            return redirect('documents:document_detail', pk=document.id)
        
        messages.error(request, _('Stato non valido.'))
    
    return redirect('documents:document_detail', pk=document.id)

class AddDocumentNoteView(LoginRequiredMixin, CreateView):
    """Vista per aggiungere una nota a un documento"""
    model = DocumentNote
    form_class = DocumentNoteForm
    template_name = 'documents/document_note_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        document_id = self.kwargs.get('pk')
        context['document'] = get_object_or_404(ProjectDocument, id=document_id)
        return context
    
    def form_valid(self, form):
        document_id = self.kwargs.get('pk')
        document = get_object_or_404(ProjectDocument, id=document_id)
        
        form.instance.document = document
        form.instance.created_by = self.request.user
        
        response = super().form_valid(form)
        messages.success(self.request, _('Nota aggiunta con successo.'))
        return response
    
    def get_success_url(self):
        return reverse('documents:document_detail', kwargs={'pk': self.kwargs.get('pk')})
    

class AddDocumentTaskView(LoginRequiredMixin, CreateView):
    """Vista per aggiungere un'attività a un documento"""
    model = DocumentTask
    form_class = DocumentTaskForm
    template_name = 'documents/document_task_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        document_id = self.kwargs.get('pk')
        context['document'] = get_object_or_404(ProjectDocument, id=document_id)
        return context
    
    def form_valid(self, form):
        document_id = self.kwargs.get('pk')
        document = get_object_or_404(ProjectDocument, id=document_id)
        
        form.instance.document = document
        form.instance.created_by = self.request.user
        
        response = super().form_valid(form)
        messages.success(self.request, _('Attività aggiunta con successo.'))
        return response
    
    def get_success_url(self):
        return reverse('documents:document_detail', kwargs={'pk': self.kwargs.get('pk')})
    
class UpdateDocumentTaskView(LoginRequiredMixin, UpdateView):
    """Vista per aggiornare lo stato di un'attività legata a un documento"""
    model = DocumentTask
    fields = ['status']
    template_name = 'documents/document_task_update_form.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['document'] = self.object.document
        return context
    
    def form_valid(self, form):
        # Se il nuovo stato è "completato", registra la data di completamento
        if form.instance.status == 'completed' and not form.instance.completed_at:
            form.instance.completed_at = timezone.now()
            
        response = super().form_valid(form)
        messages.success(self.request, _('Stato attività aggiornato con successo.'))
        return response
    
    def get_success_url(self):
        return reverse('documents:document_detail', kwargs={'pk': self.object.document.id})    

class DocumentCategoryListView(LoginRequiredMixin, ListView):
    """Vista per elencare le categorie di documenti"""
    model = DocumentCategory
    template_name = 'documents/document_category_list.html'
    context_object_name = 'categories'
    
    def get_queryset(self):
        return DocumentCategory.objects.all().order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Aggiungi il conteggio dei documenti per ogni categoria
        for category in context['categories']:
            category.document_count = Document.objects.filter(category=category).count()
        return context

class DocumentCategoryCreateView(LoginRequiredMixin, CreateView):
    """Vista per creare una nuova categoria di documenti"""
    model = DocumentCategory
    template_name = 'documents/document_category_form.html'
    form_class = DocumentCategoryForm
    success_url = reverse_lazy('documents:category_list')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Categoria creata con successo.'))
        return response

class DocumentCategoryUpdateView(LoginRequiredMixin, UpdateView):
    """Vista per aggiornare una categoria di documenti esistente"""
    model = DocumentCategory
    template_name = 'documents/document_category_form.html'
    form_class = DocumentCategoryForm
    success_url = reverse_lazy('documents:category_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['is_update'] = True
        return context
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, _('Categoria aggiornata con successo.'))
        return response

class DocumentCategoryDeleteView(LoginRequiredMixin, DeleteView):
    """Vista per eliminare una categoria di documenti"""
    model = DocumentCategory
    template_name = 'documents/document_category_confirm_delete.html'
    success_url = reverse_lazy('documents:category_list')
    context_object_name = 'category'
    
    def delete(self, request, *args, **kwargs):
        category = self.get_object()
        try:
            # Controlliamo se la categoria è usata
            if Document.objects.filter(category=category).exists():
                messages.error(request, _('Non è possibile eliminare una categoria utilizzata da documenti esistenti.'))
                return redirect('documents:category_list')
                
            response = super().delete(request, *args, **kwargs)
            messages.success(request, _('Categoria eliminata con successo.'))
            return response
        except Exception as e:
            messages.error(request, _('Errore durante l\'eliminazione: {}').format(str(e)))
            return redirect('documents:category_list')